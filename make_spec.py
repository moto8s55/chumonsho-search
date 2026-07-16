# -*- coding: utf-8 -*-
"""
書籍仕様書 検索ページ生成スクリプト

Excel（商品コード / ISBN / 書名 / オビ / カバー / …）を読み込み、
商品コード・ISBN・書名で検索できる仕様書検索ページ spec-search.html を生成します。

使い方:
    pip install openpyxl
    python make_spec.py

その後 git add/commit/push すれば GitHub Pages に反映され、Notionに埋め込めます。
"""

import json
import unicodedata
from pathlib import Path

import openpyxl

# ===================== 設定 =====================
# 仕様書Excelの場所。ファイルを直接指定してもよいし、フォルダを指定すれば
# その中の .xlsx を自動で探します（ファイル名が変わっても動くように）。
EXCEL = r"C:\Users\moto8\OneDrive\デスクトップ\興陽館共有フォルダ★★★2026_02_20\3 営業共有\営業\旧PCデスクトップ\specification document"

# 検索対象の列（この3つが検索ワードになる）
SEARCH_COLS = ["商品コード", "ISBN", "書名"]
# 仕様として表示する列（空欄は自動で省略）
SPEC_COLS = ["オビ", "カバー", "表紙", "別丁扉", "本文", "スリップ", "口絵", "ジャバラ"]
# ================================================


def norm(s: str) -> str:
    """全角半角・大小・カタカナ差を吸収（日本語・英数どちらでも検索できるように）。"""
    s = unicodedata.normalize("NFKC", str(s)).lower()
    out = []
    for ch in s:
        code = ord(ch)
        if 0x30A1 <= code <= 0x30F6:  # カタカナ→ひらがな
            out.append(chr(code - 0x60))
        else:
            out.append(ch)
    s = "".join(out)
    for c in [" ", "　", "\t", "_", "-", "・", "／", "/", "\\", "＜", "＞", "<", ">"]:
        s = s.replace(c, "")
    return s


def cell(v):
    return "" if v is None else str(v).strip()


def resolve_excel(path):
    """ファイル指定ならそのまま。フォルダ指定ならその中の .xlsx を自動選択。"""
    p = Path(path)
    if p.is_file():
        return p
    folder = p if p.is_dir() else p.parent
    if not folder.exists():
        raise SystemExit(f"[エラー] フォルダが見つかりません: {folder}")
    # 一時ファイル(~$)を除く .xlsx。名前に spec を含むものを優先、無ければ更新が新しい順。
    xlsx = [f for f in folder.glob("*.xlsx") if not f.name.startswith("~$")]
    if not xlsx:
        raise SystemExit(f"[エラー] .xlsx が見つかりません: {folder}")
    xlsx.sort(key=lambda f: (0 if "spec" in f.name.lower() else 1, -f.stat().st_mtime))
    return xlsx[0]


def build(excel_path=EXCEL, out_path=None):
    excel_path = resolve_excel(excel_path)
    print(f"[Excel] {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    headers = [cell(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    idx = {h: i for i, h in enumerate(headers)}

    books = []
    for r in range(2, ws.max_row + 1):
        vals = [cell(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        code = vals[idx["商品コード"]] if "商品コード" in idx else ""
        isbn = vals[idx["ISBN"]] if "ISBN" in idx else ""
        title = vals[idx["書名"]] if "書名" in idx else ""
        if not (code or isbn or title):
            continue  # 完全空行はスキップ
        specs = []
        for col in SPEC_COLS:
            if col in idx:
                val = vals[idx[col]]
                if val:
                    specs.append([col, val])
        key = norm(code) + "\n" + norm(isbn) + "\n" + norm(title)
        books.append(
            {"code": code, "isbn": isbn, "title": title, "specs": specs, "key": key}
        )

    # 刊行が古い順（Excelの並び）を維持
    out_dir = Path(__file__).resolve().parent
    out = Path(out_path) if out_path else (out_dir / "spec-search.html")
    inner = json.dumps(books, ensure_ascii=False)[1:-1]
    html = TEMPLATE.replace("/*__DATA__*/", inner)
    out.write_text(html, encoding="utf-8")
    print(f"[完了] {len(books)} 冊を索引化しました → {out}")
    return len(books)


TEMPLATE = r"""<!DOCTYPE html>
<html lang="ja">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>書籍仕様書 検索</title>
<style>
  :root { --bg:#111; --panel:#1b1b1b; --line:#2c2c2c; --txt:#e8e8e8; --dim:#8a8a8a; --accent:#2ee6a0; }
  * { box-sizing:border-box; }
  body { margin:0; font-family:-apple-system,"Segoe UI","Hiragino Sans","Noto Sans JP",sans-serif;
         background:var(--bg); color:var(--txt); padding:28px 16px; }
  .wrap { max-width:860px; margin:0 auto; }
  h1 { text-align:center; letter-spacing:.16em; color:var(--accent);
       font-size:22px; font-weight:800; margin:0 0 4px; font-family:ui-monospace,Menlo,Consolas,monospace; }
  .sub { text-align:center; color:var(--dim); font-size:12px; margin-bottom:22px; }
  .searchbar { display:flex; gap:10px; }
  .field { position:relative; flex:1; }
  #q { width:100%; background:#0d0d0d; border:1px solid var(--line); color:var(--txt);
       font-size:18px; padding:14px 42px 14px 16px; border-radius:8px; outline:none; }
  #q:focus { border-color:var(--accent); }
  #clear { position:absolute; right:8px; top:50%; transform:translateY(-50%);
           background:#333; color:#ddd; border:none; width:26px; height:26px; border-radius:50%;
           font-size:15px; cursor:pointer; line-height:1; display:none; }
  #clear:hover { background:var(--accent); color:#062; }
  .hint { text-align:center; color:var(--dim); font-size:11px; margin:8px 0 18px; }
  .count { color:var(--dim); font-size:12px; margin:6px 2px 10px; }
  .list { list-style:none; padding:0; margin:0; }
  .item { background:var(--panel); border:1px solid var(--line); border-radius:8px;
          padding:12px 14px; margin-bottom:8px; }
  .title { font-size:16px; font-weight:700; margin-bottom:4px; }
  .ids { font-size:12px; color:var(--dim); margin-bottom:8px; display:flex; gap:16px; flex-wrap:wrap; }
  .ids b { color:#bdbdbd; font-weight:600; }
  .specs { display:grid; grid-template-columns:auto 1fr; gap:4px 12px; font-size:13px; }
  .specs .k { color:var(--accent); white-space:nowrap; }
  .specs .v { color:var(--txt); word-break:break-word; }
  .nospec { font-size:12px; color:var(--dim); font-style:italic; }
  mark { background:transparent; color:var(--accent); font-weight:800; }
  .empty { text-align:center; color:var(--dim); padding:40px 0; }
</style>
</head>
<body>
<div class="wrap">
  <h1>書籍仕様書 SEARCH</h1>
  <div class="sub">商品コード・ISBN・書名で検索（日本語／英数どちらでもOK）</div>
  <div class="searchbar">
    <div class="field">
      <input id="q" type="search" placeholder="例: 3355 / 9784877233563 / あっこです / Misa …" autocomplete="off" autofocus>
      <button id="clear" title="クリア">×</button>
    </div>
  </div>
  <div class="hint">全角半角・カタカナ/ひらがな・スペースの違いを無視してヒットします</div>
  <div class="count" id="count"></div>
  <ul class="list" id="list"></ul>
</div>

<script>
const DATA = [/*__DATA__*/];

function norm(s){
  s = (s||"").normalize("NFKC").toLowerCase();
  let out = "";
  for (const ch of s){
    const c = ch.codePointAt(0);
    if (c >= 0x30A1 && c <= 0x30F6) out += String.fromCodePoint(c - 0x60);
    else out += ch;
  }
  return out.replace(/[\s_\-・／/\\＜＞<>　]/g, "");
}
function score(qn, key){
  if (!qn) return 0.01;
  const idx = key.indexOf(qn);
  if (idx >= 0) return 1000 - idx;
  let i = 0;
  for (const ch of key){ if (ch === qn[i]) i++; if (i === qn.length) break; }
  if (i === qn.length) return 400 - key.length * 0.1;
  const grams = q => { const g=[]; for(let k=0;k<q.length-1;k++) g.push(q.slice(k,k+2)); return g; };
  const qg = grams(qn); if (!qg.length) return 0;
  let hit = 0; for (const g of qg) if (key.includes(g)) hit++;
  const ratio = hit / qg.length;
  return ratio >= 0.5 ? 100 * ratio : 0;
}
function escapeHtml(s){ return (s==null?"":String(s)).replace(/[&<>"']/g, m => ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[m])); }
function hl(text, qn){
  if (!qn) return escapeHtml(text);
  const raw = String(text).toLowerCase();
  const p = raw.indexOf(qn);
  if (p < 0) return escapeHtml(text);
  return escapeHtml(text.slice(0,p)) + "<mark>" + escapeHtml(text.slice(p,p+qn.length)) + "</mark>" + escapeHtml(text.slice(p+qn.length));
}

const listEl = document.getElementById("list");
const countEl = document.getElementById("count");
const qEl = document.getElementById("q");
const clearBtn = document.getElementById("clear");

function render(items, qn){
  if (!items.length){ listEl.innerHTML = '<div class="empty">該当なし</div>'; return; }
  listEl.innerHTML = items.map(b => {
    const specs = b.specs.length
      ? '<div class="specs">' + b.specs.map(s => `<div class="k">${escapeHtml(s[0])}</div><div class="v">${escapeHtml(s[1])}</div>`).join("") + '</div>'
      : '<div class="nospec">（仕様データなし）</div>';
    return `<li class="item">
      <div class="title">${hl(b.title, qn)}</div>
      <div class="ids"><span><b>商品コード</b> ${hl(b.code, qn)}</span><span><b>ISBN</b> ${hl(b.isbn, qn)}</span></div>
      ${specs}
    </li>`;
  }).join("");
}

function run(){
  const qn = norm(qEl.value);
  clearBtn.style.display = qEl.value ? "block" : "none";
  let items;
  if (!qn){
    items = DATA.slice();
    countEl.textContent = `全 ${DATA.length} 冊（刊行が古い順）`;
  } else {
    items = DATA.map(b => ({b, s: score(qn, b.key)}))
                .filter(x => x.s > 0)
                .sort((a,b) => b.s - a.s)
                .map(x => x.b);
    countEl.textContent = `${items.length} 件ヒット`;
  }
  render(items, qn);
}

qEl.addEventListener("input", run);
clearBtn.addEventListener("click", () => { qEl.value = ""; qEl.focus(); run(); });
run();
</script>
</body>
</html>
"""


if __name__ == "__main__":
    import sys
    build(*(sys.argv[1:3]))
